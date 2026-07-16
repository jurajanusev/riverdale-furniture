import csv
import io
import json
import os
import re
import secrets
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlparse

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer

from database import (
    architect_choice_counts, architect_choice_ids, create_share_link, delete_product,
    distinct_values, get_product, get_share_link, init_db, list_products,
    list_share_links, move_product, revoke_share_link, save_product, set_architect_choice, update_product,
)
from riverdale_catalog import CATEGORIES, CATEGORY_BY_ID, DEFAULT_CONTEXT, SPACES, SPACE_BY_ID, validate_context
from scrapers import SCRAPERS, search_all
from scrapers.base import browser_profile_dir
from models import Product
from search_worker import run_search


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("RIVERDALE_DATA_DIR", BASE_DIR / "data"))
UPLOAD_DIR = Path(os.environ.get("RIVERDALE_UPLOAD_DIR", BASE_DIR / "static" / "uploads"))
EXPORT_DIR = Path(os.environ.get("RIVERDALE_EXPORT_DIR", BASE_DIR / "exports"))
SEARCH_JOB_DIR = DATA_DIR / "search_jobs"
SEARCH_EXECUTOR = ThreadPoolExecutor(max_workers=1, thread_name_prefix="riverdale-search")
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
ALLOWED_STORES = [
    "IKEA Slovensko", "IKEA Rakúsko", "Möbelix Slovensko", "Möbelix Rakúsko",
    "XXXLutz Slovensko", "XXXLutz Rakúsko", "Mömax Rakúsko", "Sconto Slovensko",
    "JYSK Slovensko", "JYSK Rakúsko", "Bonami Slovensko", "ASKO Slovensko",
    "FAVI Slovensko",
]
STATUS_LABELS = {"unreviewed": "Neposúdené", "approved": "Schválené", "rejected": "Vyradené", "maybe": "Možno"}
CAPTCHA_STORES = {
    "moebelix-sk": "Möbelix Slovensko", "moebelix-at": "Möbelix Rakúsko",
    "xxxlutz-sk": "XXXLutz Slovensko", "xxxlutz-at": "XXXLutz Rakúsko",
    "moemax-at": "Mömax Rakúsko", "sconto-sk": "Sconto Slovensko",
}
SEARCH_FORM_FIELDS = (
    "search_min_price", "search_max_price", "search_color", "search_material",
    "search_in_stock", "search_min_width", "search_max_width", "search_max_depth",
    "search_max_height", "search_bed_width", "search_bed_length",
    "search_slats_included", "search_mattress_included",
)


def search_form_values(values):
    return {key: str(values.get(key, "")).strip() for key in SEARCH_FORM_FIELDS}


def search_criteria(values):
    context = validate_context(values)
    numeric_fields = {
        "search_min_price": "min_price", "search_max_price": "max_price",
        "search_min_width": "min_width", "search_max_width": "max_width",
        "search_max_depth": "max_depth", "search_max_height": "max_height",
    }
    numbers = {}
    for form_key, criteria_key in numeric_fields.items():
        value = str(values.get(form_key, "")).strip()
        if not value:
            numbers[criteria_key] = ""
            continue
        try:
            number = float(value.replace(",", "."))
            if number < 0:
                raise ValueError
            numbers[criteria_key] = number
        except ValueError as exc:
            raise ValueError("Cena a rozmery musia byť nezáporné čísla.") from exc
    if numbers["min_price"] != "" and numbers["max_price"] != "" and numbers["min_price"] > numbers["max_price"]:
        raise ValueError("Minimálna cena nemôže byť vyššia ako maximálna cena.")
    return {
        **context, **numbers,
        "color": str(values.get("search_color", "")).strip(),
        "material": str(values.get("search_material", "")).strip(),
        "in_stock": values.get("search_in_stock") == "yes",
        "bed_width": str(values.get("search_bed_width", "")).strip(),
        "bed_length": str(values.get("search_bed_length", "")).strip(),
        "slats_included": values.get("search_slats_included", "any"),
        "mattress_included": values.get("search_mattress_included", "any"),
    }


def read_search_job(job_id):
    if not re.fullmatch(r"[A-Za-z0-9_-]{12,64}", job_id or ""):
        return None
    try:
        return json.loads((SEARCH_JOB_DIR / f"{job_id}.json").read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None


def create_app(test_config=None):
    app = Flask(__name__)
    app.config.update(SECRET_KEY=os.environ.get("RIVERDALE_SECRET", secrets.token_hex(16)), MAX_CONTENT_LENGTH=8 * 1024 * 1024)
    if test_config:
        app.config.update(test_config)
    collector_signer = URLSafeTimedSerializer(app.config["SECRET_KEY"], salt="riverdale-collector")
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    SEARCH_JOB_DIR.mkdir(parents=True, exist_ok=True)
    init_db()

    if not app.config.get("TESTING"):
        for pending_path in SEARCH_JOB_DIR.glob("*.json"):
            try:
                pending = json.loads(pending_path.read_text(encoding="utf-8"))
            except (OSError, ValueError):
                continue
            if pending.get("state") not in {"queued", "running"}:
                continue
            criteria = pending.get("criteria")
            if isinstance(criteria, dict):
                pending["state"] = "queued"
                pending_path.write_text(json.dumps(pending, ensure_ascii=False), encoding="utf-8")
                SEARCH_EXECUTOR.submit(run_search, criteria, pending_path)
            else:
                pending.update(state="error", error="Stará úloha bola prerušená. Spustite vyhľadávanie znova.")
                pending_path.write_text(json.dumps(pending, ensure_ascii=False), encoding="utf-8")

    def is_cloud_runtime():
        return bool(os.environ.get("RENDER") or os.environ.get("RENDER_SERVICE_ID"))

    def start_all_verifications(values):
        criteria = search_criteria(values)
        for key in ("collector_token", "collector_cloud_url"):
            if values.get(key):
                criteria[key] = values.get(key)
        jobs = []
        for store_name in CAPTCHA_STORES.values():
            scraper_class = next((cls for cls in SCRAPERS if cls.store == store_name), None)
            if not scraper_class:
                continue
            target_url = scraper_class(criteria=criteria).catalog_url_for_search()
            if target_url:
                jobs.append({"store": store_name, "url": target_url})
        if not jobs:
            raise ValueError("Pre túto kategóriu nie je pripravený žiadny blokovaný obchod.")
        subprocess.Popen(
            [
                sys.executable, str(BASE_DIR / "verify_all_stores.py"),
                json.dumps(jobs, ensure_ascii=False), json.dumps(criteria, ensure_ascii=False),
            ],
            cwd=BASE_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        return criteria, len(jobs)

    @app.before_request
    def require_admin_login():
        password = os.environ.get("RIVERDALE_ADMIN_PASSWORD", "")
        if not password or session.get("riverdale_admin"):
            return None
        public_endpoint = request.endpoint in {"login", "healthcheck", "static", "uploaded_image", "collector_import"}
        public_path = request.path.startswith("/architect/")
        if public_endpoint or public_path:
            return None
        return redirect(url_for("login", next=request.full_path.rstrip("?")))

    @app.route("/login", methods=["GET", "POST"])
    def login():
        password = os.environ.get("RIVERDALE_ADMIN_PASSWORD", "")
        if not password:
            return redirect(url_for("index"))
        if request.method == "POST":
            if secrets.compare_digest(request.form.get("password", ""), password):
                session["riverdale_admin"] = True
                destination = request.form.get("next", "")
                if not destination.startswith("/") or destination.startswith("//"):
                    destination = url_for("index")
                return redirect(destination)
            flash("Nesprávne heslo.", "danger")
        return render_template("login.html", next=request.args.get("next", request.form.get("next", "")))

    @app.post("/logout")
    def logout():
        session.clear()
        return redirect(url_for("login"))

    @app.get("/healthz")
    def healthcheck():
        return jsonify(ok=True)

    @app.get("/api/search-jobs/<job_id>")
    def search_job_status(job_id):
        status = read_search_job(job_id)
        if not status:
            return jsonify(error="Vyhľadávanie neexistuje alebo bolo prerušené."), 404
        return jsonify(status)

    @app.get("/media/<path:filename>")
    def uploaded_image(filename):
        return send_from_directory(UPLOAD_DIR, Path(filename).name)

    @app.post("/api/collector/products")
    def collector_import():
        expected_token = os.environ.get("RIVERDALE_SYNC_TOKEN") or os.environ.get("RIVERDALE_ADMIN_PASSWORD", "")
        authorization = request.headers.get("Authorization", "")
        supplied_token = authorization[7:] if authorization.startswith("Bearer ") else ""
        static_valid = bool(expected_token and supplied_token and secrets.compare_digest(supplied_token, expected_token))
        signed_valid = False
        if supplied_token and not static_valid:
            try:
                signed_valid = collector_signer.loads(supplied_token, max_age=7200).get("purpose") == "collector"
            except (BadSignature, SignatureExpired, AttributeError):
                signed_valid = False
        if not static_valid and not signed_valid:
            return jsonify(error="Neplatný synchronizačný kľúč."), 401
        payload = request.get_json(silent=True) or {}
        raw_products = payload.get("products")
        if not isinstance(raw_products, list) or not raw_products or len(raw_products) > 50:
            return jsonify(error="Očakáva sa 1 až 50 produktov."), 400
        scraper_by_store = {cls.store: cls for cls in SCRAPERS}
        allowed_fields = set(Product.__dataclass_fields__) - {
            "id", "internal_id", "created_at", "updated_at", "local_image",
            "approval_status", "notes",
        }
        imported, rejected = 0, []
        for position, raw in enumerate(raw_products, start=1):
            if not isinstance(raw, dict):
                rejected.append({"position": position, "reason": "Neplatný záznam"})
                continue
            store = str(raw.get("store", ""))
            product_url = str(raw.get("product_url", "")).strip()
            scraper_class = scraper_by_store.get(store)
            context = validate_context(raw)
            try:
                valid_url = bool(scraper_class and scraper_class(criteria=context).is_product_url(product_url))
            except (TypeError, ValueError):
                valid_url = False
            if not raw.get("name") or not valid_url:
                rejected.append({"position": position, "reason": "Neplatný názov, obchod alebo URL"})
                continue
            product = {key: raw[key] for key in allowed_fields if key in raw}
            product.update(context)
            product.update({
                "name": str(raw["name"]).strip()[:300],
                "store": store,
                "product_url": product_url,
                "last_checked": str(raw.get("last_checked") or date.today().isoformat()),
                "approval_status": "unreviewed",
                "local_image": "",
                "notes": "",
            })
            save_product(product)
            imported += 1
        return jsonify(ok=True, imported=imported, rejected=rejected), 200 if imported else 400

    @app.get("/")
    def index():
        context = validate_context(request.args)
        filters = {key: request.args.get(key, "").strip() for key in ("store", "country", "max_price", "color", "material", "availability", "approval_status")}
        filters.update({key: context[key] for key in ("space_id", "room", "main_category", "item_type")})
        sort = request.args.get("sort", "store")
        products = list_products(filters, sort)
        selection_count = len(list_products({
            "space_id": context["space_id"], "room": context["room"],
            "approval_status": "approved",
        }))
        options = {key: distinct_values(key) for key in ("store", "country", "color", "material", "availability")}
        search_job_id = request.args.get("search_job", "")
        search_job = read_search_job(search_job_id) if search_job_id else None
        captcha_statuses = {}
        for key, store in CAPTCHA_STORES.items():
            status_path = browser_profile_dir(store) / "riverdale-status.json"
            try:
                captcha_statuses[key] = json.loads(status_path.read_text(encoding="utf-8"))
            except (OSError, ValueError):
                captcha_statuses[key] = {}
        return render_template(
            "index.html", products=products, filters=filters, context=context, sort=sort,
            options=options, stores=ALLOWED_STORES, status_labels=STATUS_LABELS,
            spaces=SPACES, categories=CATEGORIES, space_by_id=SPACE_BY_ID,
            category_by_id=CATEGORY_BY_ID, selection_count=selection_count,
            captcha_stores=CAPTCHA_STORES, captcha_statuses=captcha_statuses,
            cloud_runtime=is_cloud_runtime(),
            search_values=search_form_values(request.args),
            collector_token=collector_signer.dumps({"purpose": "collector"}) if is_cloud_runtime() else "",
            search_job_id=search_job_id, search_job=search_job,
        )

    @app.get("/selection")
    def selection():
        context = validate_context(request.args)
        products = list_products({
            "space_id": context["space_id"], "room": context["room"],
            "approval_status": "approved",
        }, "name")
        products.sort(key=lambda product: (
            CATEGORY_BY_ID.get(product.main_category, {}).get("name", product.main_category),
            product.item_type, product.name,
        ))
        groups = []
        for product in products:
            key = (product.main_category, product.item_type)
            if not groups or groups[-1]["key"] != key:
                groups.append({
                    "key": key,
                    "category_name": CATEGORY_BY_ID.get(product.main_category, {}).get("name", product.main_category),
                    "item_type": product.item_type,
                    "products": [],
                })
            groups[-1]["products"].append(product)
        total_price = sum(product.frame_price or 0 for product in products)
        share_links = list_share_links(context["space_id"], context["room"])
        architect_counts = architect_choice_counts(context["space_id"], context["room"])
        return render_template(
            "selection.html", context=context, products=products, groups=groups,
            total_price=total_price, spaces=SPACES, categories=CATEGORIES,
            share_links=share_links, architect_counts=architect_counts,
        )

    @app.post("/shares")
    def create_architect_share():
        context = validate_context(request.form)
        architect_name = request.form.get("architect_name", "").strip()[:80] or "Architekt"
        token = secrets.token_urlsafe(24)
        expires_at = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat(timespec="seconds")
        create_share_link(token, context["space_id"], context["room"], architect_name, expires_at)
        flash("Odkaz pre architekta bol vytvorený a platí 30 dní.", "success")
        return redirect(url_for(
            "selection", space_id=context["space_id"], room=context["room"],
            main_category=context["main_category"], item_type=context["item_type"],
        ))

    @app.post("/shares/<token>/revoke")
    def revoke_architect_share(token):
        share = get_share_link(token, active_only=False)
        if not share:
            return jsonify(error="Odkaz neexistuje."), 404
        revoke_share_link(token)
        context = validate_context({"space_id": share["space_id"], "room": share["room"]})
        flash("Odkaz pre architekta bol zrušený.", "info")
        return redirect(url_for("selection", space_id=context["space_id"], room=context["room"]))

    @app.get("/architect/<token>")
    def architect_review(token):
        share = get_share_link(token)
        if not share:
            return render_template("architect_unavailable.html"), 410
        products = list_products({
            "space_id": share["space_id"], "room": share["room"],
            "approval_status": "approved",
        }, "name")
        products.sort(key=lambda product: (
            CATEGORY_BY_ID.get(product.main_category, {}).get("name", product.main_category),
            product.item_type, product.name,
        ))
        groups = []
        for product in products:
            key = (product.main_category, product.item_type)
            if not groups or groups[-1]["key"] != key:
                groups.append({
                    "key": key,
                    "category_name": CATEGORY_BY_ID.get(product.main_category, {}).get("name", product.main_category),
                    "item_type": product.item_type,
                    "products": [],
                })
            groups[-1]["products"].append(product)
        selected_ids = architect_choice_ids(share["id"])
        return render_template(
            "architect.html", share=share, groups=groups, products=products,
            selected_ids=selected_ids,
            space_name=SPACE_BY_ID.get(share["space_id"], {}).get("name", share["space_id"]),
        )

    @app.post("/architect/<token>/choice/<int:product_id>")
    def architect_choice(token, product_id):
        share = get_share_link(token)
        if not share:
            return jsonify(error="Odkaz už nie je platný."), 410
        product = get_product(product_id)
        if not product or product.approval_status != "approved" or product.space_id != share["space_id"] or product.room != share["room"]:
            return jsonify(error="Produkt nie je súčasťou tohto výberu."), 404
        data = request.get_json(silent=True) or {}
        selected = data.get("selected") is True
        set_architect_choice(share["id"], product_id, selected)
        return jsonify(ok=True, selected=selected)

    @app.get("/architect/<token>/image/<int:product_id>")
    def architect_product_image(token, product_id):
        share = get_share_link(token)
        product = get_product(product_id)
        if not share or not product or not product.local_image or product.space_id != share["space_id"] or product.room != share["room"]:
            return jsonify(error="Obrázok nie je dostupný."), 404
        image_path = (UPLOAD_DIR / Path(product.local_image).name).resolve()
        if UPLOAD_DIR.resolve() not in image_path.parents or not image_path.is_file():
            return jsonify(error="Obrázok nie je dostupný."), 404
        return send_file(image_path)

    @app.post("/products")
    def add_product():
        context = validate_context(request.form)
        try:
            product = manual_product_from_request(request)
            save_product(product)
            flash("Produkt bol pridaný.", "success")
        except ValueError as exc:
            flash(str(exc), "danger")
        return redirect(url_for("index", **{key: context[key] for key in ("space_id", "room", "main_category", "item_type")}))

    @app.patch("/api/products/<int:product_id>")
    def change_product(product_id):
        product = get_product(product_id)
        if not product:
            return jsonify(error="Produkt neexistuje."), 404
        data = request.get_json(silent=True) or {}
        status = data.get("approval_status")
        if status is not None and status not in STATUS_LABELS:
            return jsonify(error="Neplatný stav."), 400
        destination_space_id = data.get("space_id")
        destination_room = str(data.get("room", "")).strip()
        moved = False
        if destination_space_id is not None or destination_room:
            destination = SPACE_BY_ID.get(str(destination_space_id or ""))
            if not destination or destination_room not in destination["rooms"]:
                return jsonify(error="Neplatný cieľový priestor alebo miestnosť."), 400
            move_result = move_product(
                product_id, destination["id"], destination["name"], destination_room,
            )
            if move_result == "duplicate":
                return jsonify(error="Tento produkt už v cieľovej miestnosti existuje."), 409
            if move_result == "missing":
                return jsonify(error="Produkt neexistuje."), 404
            moved = True
        update_product(product_id, approval_status=status, notes=data.get("notes"))
        product = get_product(product_id)
        return jsonify(
            ok=True, moved=moved, approval_status=product.approval_status,
            label=STATUS_LABELS.get(product.approval_status, ""),
            selection_url=url_for(
                "selection", space_id=product.space_id, room=product.room,
                main_category=product.main_category, item_type=product.item_type,
            ),
            destination=f"{product.space_name} · {product.room}",
        )

    @app.delete("/api/products/<int:product_id>")
    def remove_product(product_id):
        if not delete_product(product_id):
            return jsonify(error="Produkt neexistuje."), 404
        return jsonify(ok=True)

    @app.post("/search")
    def search_products():
        context = validate_context(request.form)
        try:
            criteria = search_criteria(request.form)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("index", **context, **search_form_values(request.form)))
        job_id = secrets.token_urlsafe(16)
        job_path = SEARCH_JOB_DIR / f"{job_id}.json"
        job_path.write_text(
            json.dumps({"state": "queued", "messages": [], "imported": 0, "criteria": criteria}, ensure_ascii=False),
            encoding="utf-8",
        )
        try:
            SEARCH_EXECUTOR.submit(run_search, criteria, job_path)
        except RuntimeError as exc:
            job_path.write_text(
                json.dumps({"state": "error", "messages": [], "imported": 0, "error": str(exc)}, ensure_ascii=False),
                encoding="utf-8",
            )
        return redirect(url_for("index", **context, **search_form_values(request.form), search_job=job_id))

    @app.post("/verify-store/<store_key>")
    def verify_store(store_key):
        context = validate_context(request.form)
        if is_cloud_runtime():
            flash(
                "CAPTCHA sa musí overiť na vašom počítači. Spustite start_collector.ps1 "
                "a ručné overenie použite v lokálnej aplikácii na http://127.0.0.1:5000.",
                "info",
            )
            return redirect(url_for("index", **context))
        store_name = CAPTCHA_STORES.get(store_key)
        scraper_class = next((cls for cls in SCRAPERS if cls.store == store_name), None)
        if not scraper_class:
            flash("Neznámy obchod pre ručné overenie.", "danger")
            return redirect(url_for("index", **context))
        try:
            criteria = search_criteria(request.form)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("index", **context, **search_form_values(request.form)))
        scraper = scraper_class(criteria=criteria)
        target_url = scraper.catalog_url_for_search()
        if not target_url:
            flash("Pre túto kategóriu obchod nemá pripravenú cieľovú stránku.", "danger")
            return redirect(url_for("index", **context))
        try:
            subprocess.Popen(
                [
                    sys.executable, str(BASE_DIR / "verify_store.py"), store_name,
                    target_url, context["item_type"], json.dumps(criteria, ensure_ascii=False),
                ],
                cwd=BASE_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
        except OSError as exc:
            flash(f"Prehliadač sa nepodarilo spustiť: {exc}", "danger")
            return redirect(url_for("index", **context))
        flash(
            f"Otvorilo sa ručné overenie pre {store_name}. Vyriešte CAPTCHA a okno nechajte otvorené; po dokončení sa produkty automaticky odošlú do cloudu, ak bol lokálny zberač spustený cez start_collector.ps1.",
            "info",
        )
        return redirect(url_for("index", **context, **search_form_values(request.form)))

    @app.get("/local-verify-store/<store_key>")
    def local_verify_store(store_key):
        if is_cloud_runtime() or request.remote_addr not in {"127.0.0.1", "::1"}:
            return "Lokálne overenie nie je dostupné.", 403
        context = validate_context(request.args)
        store_name = CAPTCHA_STORES.get(store_key)
        scraper_class = next((cls for cls in SCRAPERS if cls.store == store_name), None)
        if not scraper_class:
            return "Neznámy obchod.", 404
        try:
            criteria = search_criteria(request.args)
        except ValueError as exc:
            return str(exc), 400
        scraper = scraper_class(criteria=criteria)
        collector_token = request.args.get("collector_token", "").strip()
        collector_cloud_url = request.args.get("collector_cloud_url", "").strip().rstrip("/")
        cloud_parts = urlparse(collector_cloud_url)
        if not collector_token or cloud_parts.scheme != "https" or cloud_parts.hostname != "riverdale-furniture.onrender.com":
            return "Neplatné cloudové prepojenie. Obnovte cloudovú stránku.", 400
        criteria["collector_token"] = collector_token
        criteria["collector_cloud_url"] = collector_cloud_url
        target_url = scraper.catalog_url_for_search()
        if not target_url:
            return "Pre túto kategóriu obchod nemá cieľovú stránku.", 400
        try:
            subprocess.Popen(
                [
                    sys.executable, str(BASE_DIR / "verify_store.py"), store_name,
                    target_url, context["item_type"], json.dumps(criteria, ensure_ascii=False),
                ],
                cwd=BASE_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
        except OSError as exc:
            return f"Prehliadač sa nepodarilo spustiť: {exc}", 500
        return (
            "<!doctype html><html lang='sk'><meta charset='utf-8'>"
            f"<title>Riverdale – {store_name}</title>"
            "<body style='font:16px sans-serif;padding:24px'>"
            f"<p>Otváram CAPTCHA pre <strong>{store_name}</strong>…</p>"
            "<p>Toto pomocné okno sa automaticky zavrie.</p>"
            "<script>setTimeout(() => window.close(), 800);</script></body></html>"
        )

    @app.route("/local-verify-all", methods=["GET", "POST"])
    def local_verify_all():
        if is_cloud_runtime() or request.remote_addr not in {"127.0.0.1", "::1"}:
            return "Lokálne overenie nie je dostupné.", 403
        values = request.args if request.method == "GET" else request.form
        context = validate_context(values)
        criteria_values = dict(values)
        if request.method == "GET":
            collector_token = values.get("collector_token", "").strip()
            collector_cloud_url = values.get("collector_cloud_url", "").strip().rstrip("/")
            cloud_parts = urlparse(collector_cloud_url)
            if not collector_token or cloud_parts.scheme != "https" or cloud_parts.hostname != "riverdale-furniture.onrender.com":
                return "Neplatné cloudové prepojenie. Obnovte cloudovú stránku.", 400
            criteria_values["collector_token"] = collector_token
            criteria_values["collector_cloud_url"] = collector_cloud_url
        try:
            _, count = start_all_verifications(criteria_values)
        except (ValueError, OSError) as exc:
            if request.method == "POST":
                flash(str(exc), "danger")
                return redirect(url_for("index", **context, **search_form_values(values)))
            return str(exc), 400
        if request.method == "POST":
            flash(f"Spustilo sa postupné overenie {count} obchodov.", "info")
            return redirect(url_for("index", **context, **search_form_values(values)))
        return (
            "<!doctype html><html lang='sk'><meta charset='utf-8'>"
            "<title>Riverdale – CAPTCHA</title>"
            "<body style='font:16px sans-serif;padding:24px'>"
            f"<p>Spúšťam postupné overenie <strong>{count} obchodov</strong>…</p>"
            "<p>Po dokončení jedného obchodu sa otvorí ďalší. Toto okno sa zavrie.</p>"
            "<script>setTimeout(() => window.close(), 800);</script></body></html>"
        )

    @app.post("/refresh")
    def refresh_prices():
        refreshed, failed = 0, 0
        scraper_by_store = {cls.store: cls for cls in SCRAPERS}
        for product in list_products():
            cls = scraper_by_store.get(product.store)
            if not cls or not product.product_url:
                continue
            try:
                scraper = cls(criteria={
                    "space_id": product.space_id, "space_name": product.space_name,
                    "room": product.room, "main_category": product.main_category,
                    "item_type": product.item_type,
                })
                updated = scraper.parse_product(product.product_url, scraper.fetch(product.product_url))
                if updated:
                    updated["notes"] = product.notes
                    save_product(updated)
                    refreshed += 1
                else:
                    failed += 1
            except Exception:
                failed += 1
        flash(f"Aktualizované: {refreshed}; neoverené alebo nedostupné: {failed}.", "info")
        return redirect(url_for("index"))

    @app.get("/export/<kind>")
    def export(kind):
        products = list_products({"approval_status": "approved"}, "store")
        if kind == "xlsx":
            return export_xlsx(products)
        if kind in {"csv", "canva"}:
            return export_csv(products, canva=kind == "canva")
        return jsonify(error="Neznámy formát exportu."), 404

    return app


def optional_bool(value):
    return {"yes": True, "no": False, "unknown": None, "true": True, "false": False}.get(str(value).lower())


def manual_product_from_request(req):
    form = req.form
    context = validate_context(form)
    required = {"name": "Názov", "store": "Obchod", "price": "Cena", "product_url": "URL produktu"}
    missing = [label for key, label in required.items() if not form.get(key, "").strip()]
    if missing:
        raise ValueError("Chýbajú povinné polia: " + ", ".join(missing))
    if form["store"] not in ALLOWED_STORES:
        raise ValueError("Vybraný obchod nie je povolený.")
    parsed = urlparse(form["product_url"])
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("URL produktu musí byť platný odkaz http/https.")
    try:
        price = float(form["price"].replace(",", "."))
    except ValueError as exc:
        raise ValueError("Cena musí byť číslo.") from exc
    if price < 0:
        raise ValueError("Cena produktu nemôže byť záporná.")
    dimensions = form.get("dimensions", "").strip() or "Neoverené"
    size_match = re.search(r"(?<!\d)(\d{2,3})\s*[x×/]\s*(\d{2,3})(?!\d)", dimensions)
    width, length = (int(size_match.group(1)), int(size_match.group(2))) if size_match else (None, None)
    image_url = form.get("image_url", "").strip()
    image = req.files.get("image_file")
    local_image = ""
    if image and image.filename:
        suffix = Path(image.filename).suffix.lower()
        if suffix not in ALLOWED_IMAGE_EXTENSIONS:
            raise ValueError("Fotografia musí byť JPG, PNG alebo WebP.")
        filename = f"{secrets.token_hex(12)}{suffix}"
        image.save(UPLOAD_DIR / filename)
        local_image = f"uploads/{filename}"
    text = " ".join([form.get("name", ""), form.get("color", ""), form.get("material", ""), form.get("notes", "")])
    from scrapers.base import BaseScraper
    return {
        **context,
        "name": form["name"].strip(), "store": form["store"], "country": "Slovensko" if "Slovensko" in form["store"] else "Rakúsko",
        "frame_price": price, "sale_price": price, "currency": "EUR", "mattress_width": width, "mattress_length": length,
        "total_dimensions": dimensions, "color": form.get("color", "Neoverené").strip() or "Neoverené",
        "material": form.get("material", "Neoverené").strip() or "Neoverené", "slats_included": optional_bool(form.get("slats_included")),
        "mattress_included": optional_bool(form.get("mattress_included")), "product_url": form["product_url"].strip(), "image_url": image_url,
        "local_image": local_image, "last_checked": date.today().isoformat(), "availability": form.get("availability", "Neoverené"),
        "notes": form.get("notes", "").strip(), "approval_status": "unreviewed", "style_match_score": BaseScraper(criteria=context).score_for_context(text),
        "source": form["product_url"].strip(), "verification_data": json.dumps({"source": "manual", "checked": date.today().isoformat()}, ensure_ascii=False),
    }


def product_row(product):
    return [product.image_url or product.local_image, product.internal_id, product.space_name, product.room,
            CATEGORY_BY_ID.get(product.main_category, {}).get("name", product.main_category), product.item_type,
            product.name, product.store, product.country,
            product.frame_price, f"{product.mattress_width or '?'} × {product.mattress_length or '?'} cm", product.material,
            bool_label(product.slats_included), bool_label(product.mattress_included), product.product_url, product.notes,
            STATUS_LABELS.get(product.approval_status, product.approval_status)]


def bool_label(value):
    return "Áno" if value is True else "Nie" if value is False else "Neoverené"


def export_xlsx(products):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schválené produkty"
    headers = ["Foto", "ID", "Set", "Miestnosť", "Kategória", "Typ", "Model", "Obchod", "Krajina", "Cena", "Rozmer", "Materiál", "Rošt v cene", "Matrac v cene", "Odkaz", "Poznámka", "Stav"]
    sheet.append(headers)
    for product in products:
        sheet.append(product_row(product))
    header_fill = PatternFill("solid", fgColor="304B3D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [34, 16, 20, 22, 22, 22, 30, 22, 13, 14, 17, 24, 15, 17, 44, 34, 15]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        row[9].number_format = '#,##0.00 [$€-1]'
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name=f"riverdale-schvalene-{date.today().isoformat()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def export_csv(products, canva=False):
    stream = io.StringIO(newline="")
    stream.write("\ufeff")
    writer = csv.writer(stream, lineterminator="\r\n")
    if canva:
        writer.writerow(["product_id", "set", "room", "category", "item_type", "product_name", "store", "price", "dimensions", "material", "product_url", "image_url", "notes", "approval_status"])
        for p in products:
            writer.writerow([p.internal_id, p.space_name, p.room, p.main_category, p.item_type, p.name, p.store, p.frame_price, p.total_dimensions, p.material, p.product_url, p.image_url or p.local_image, p.notes, p.approval_status])
        name = "riverdale-canva-bulk-create.csv"
    else:
        writer.writerow(["Foto URL", "ID", "Set", "Miestnosť", "Kategória", "Typ", "Model", "Obchod", "Krajina", "Cena", "Mena", "Rozmer", "Celkové rozmery", "Farba", "Materiál", "Rošt v cene", "Matrac v cene", "Odkaz", "Poznámka", "Stav", "Kontrola", "Dostupnosť"])
        for p in products:
            writer.writerow([p.image_url or p.local_image, p.internal_id, p.space_name, p.room, p.main_category, p.item_type, p.name, p.store, p.country, p.frame_price, p.currency, f"{p.mattress_width or '?'} × {p.mattress_length or '?'}", p.total_dimensions, p.color, p.material, bool_label(p.slats_included), bool_label(p.mattress_included), p.product_url, p.notes, STATUS_LABELS[p.approval_status], p.last_checked, p.availability])
        name = "riverdale-google-sheets.csv"
    data = io.BytesIO(stream.getvalue().encode("utf-8"))
    return send_file(data, as_attachment=True, download_name=name, mimetype="text/csv; charset=utf-8")


app = create_app()

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
